import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

zoom_dict = {"12": "2 км.", "13": "800 м.", "14": "400 м.", "16": "100 м."}


def create_excel_report(data, custom_time_format):
    """Создание отчета в формате Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные"

    custom_headers = [
        "Адреса",
        "Ключевые слова",
        "Снимки карты",
        "Время проверки",
        "Масштаб карты",
    ]

    for col, header in enumerate(custom_headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, entry in enumerate(data, start=2):
        for col, header in enumerate(custom_headers, start=1):
            if header == "Снимки карты":
                screenshot_link = entry.get("screenshot")
                screenshot_text = f"Снимок карты №{row - 1}"
                cell = ws.cell(row=row, column=col, value=screenshot_text)
                cell.hyperlink = screenshot_link
            elif header == "Время проверки":
                raw_time = entry.get("time")
                formatted_time = custom_time_format(raw_time)
                ws.cell(row=row, column=col, value=formatted_time)
            elif header == "Масштаб карты":
                zoom_value = entry.get("zoom")
                zoom_text = zoom_dict.get(zoom_value, "Неизвестно")
                ws.cell(row=row, column=col, value=zoom_text)
            else:
                if header == "Адреса":
                    ws.cell(row=row, column=col, value=entry.get("address"))
                elif header == "Ключевые слова":
                    ws.cell(row=row, column=col, value=entry.get("keyword"))

    for col in range(1, len(custom_headers) + 1):
        max_length = 0
        column = get_column_letter(col)

        for row in range(1, len(data) + 2):
            cell = ws[column + str(row)]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    current_time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    filename = os.path.join(reports_dir, f"report_{current_time}.xlsx")

    wb.save(filename)
    print(f"Отчет сохранен как {filename}")
